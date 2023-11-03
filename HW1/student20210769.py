#!/usr/bin/python3

from openpyxl import load_workbook
import math

wb = load_workbook(filename = 'student.xlsx')
ws = wb['Sheet1']

total_list = []
for row in range(2, ws.max_row + 1):
    mid = ws.cell(row = row, column = 3).value
    final = ws.cell(row = row, column = 4).value
    hw = ws.cell(row = row, column = 5).value
    attendance = ws.cell(row = row, column = 6).value

    # total
    total  = mid * 0.3 + final * 0.35 + hw * 0.34 + attendance
    total_list.append(total) # list에 total 값 넣음
    ws.cell(row = row, column = 7, value = total)

# 순위
rank = [sorted(total_list, reverse=True).index(i) + 1 for i in total_list]

stuNum = ws.max_row - 1

a = math.trunc(stuNum * 0.15)
n = 0
i = 1
while a > 0:        
    if (rank.count(i) <= a):
        num = list(filter(lambda x : rank[x] == i, range(len(rank))))
        a -= rank.count(i)
        n += rank.count(i)
        sum += n
        for j in num:
            rank[j] = 'A+'
    else:
        break
    i += 1

a0 = math.trunc(stuNum * 0.3) - n
while a0 > 0:
    if (rank.count(i) <= a0):
        num = list(filter(lambda x : rank[x] == i, range(len(rank))))
        a0 -= rank.count(i)
        sum += n
        for j in num:
            rank[j] = 'A0'
    else:
        break
    i += 1

n = 0
b = math.trunc(stuNum * 0.2)
while b > 0:
    if (rank.count(i) <= b):
        num = list(filter(lambda x : rank[x] == i, range(len(rank))))
        b -= rank.count(i)
        sum += n
        n += rank.count(i)
        for j in num:
            rank[j] = 'B+'
    else:
        break
    i += 1

b0 = math.trunc(stuNum * 0.4) - n
while b0 > 0:
    if (rank.count(i) <= b0):
        num = list(filter(lambda x : rank[x] == i, range(len(rank))))
        b0 -= rank.count(i)
        for j in num:
            rank[j] = 'B0'
    else:
        break
    i += 1

n = 0
c = math.trunc(stuNum * 0.15)
while c > 0:
    if (rank.count(i) <= c):
        num = list(filter(lambda x : rank[x] == i, range(len(rank))))
        c -= rank.count(i)
        n += rank.count(i)
        for j in num:
            rank[j] = 'C+'
    else:
        break
    i += 1

for i in range(len(rank)):
    num = list(filter(lambda x : rank[x] == i, range(len(rank))))
    for j in num:
        rank[j] = 'C0'
    i += 1

k = 0
for k in range(len(rank)):
    if ((ws.cell(row = k + 2, column = 6).value) == 0 or (ws.cell(row = k + 2, column = 7).value) < 40):
        rank[k] = 'F'

i = 0
for row in range(2, ws.max_row + 1):
    ws.cell(row = row, column = 8, value = rank[i])
    i += 1

wb.save(filename = 'student.xlsx')
